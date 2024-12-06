from datetime import timedelta
from io import BytesIO
from api.CustomFilter import filter_by_group
from api.models import Group, Heat, MeetEvent
from api.serializers.GroupSerializer import GroupsListSerializer
from api.utils.HelperFunctions import format_time
from django.http import HttpResponse
from drf_spectacular.utils import OpenApiResponse, extend_schema
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView


def create_excel_file_for_event_results(swim_meet_details, event_name, results_by_group):
    workbook = Workbook()

    # Colors for ranks
    gold_fill = PatternFill(start_color="FFD700",
                            end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0",
                              end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32",
                              end_color="CD7F32", fill_type="solid")

    is_first_sheet = True
    for group_name, group_results in results_by_group.items():
        if is_first_sheet:
            results_worksheet = workbook.active
            results_worksheet.title = group_name
            is_first_sheet = False
        else:
            results_worksheet = workbook.create_sheet(title=group_name)

        results_headers = ["Rank", "Athlete", "Heat Time"]
        current_row = 1

        # Swim Meet Information
        results_worksheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(results_headers))
        swim_meet_cell = results_worksheet.cell(row=1, column=1)
        swim_meet_cell.value = f"{swim_meet_details['name']}\n{swim_meet_details['date']}\n{swim_meet_details['location']}\n"
        swim_meet_cell.style = 'Title'
        swim_meet_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

        current_row += 2

        # Event RESULTS
        results_worksheet.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=len(results_headers))
        event_meet_cell = results_worksheet.cell(row=current_row, column=1)
        event_meet_cell.value = f"{event_name}\n RESULTS"
        event_meet_cell.style = 'Headline 1'
        event_meet_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        current_row += 3

        # Title
        results_worksheet.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=len(results_headers)
        )
        group_cell = results_worksheet.cell(row=current_row, column=1)
        group_cell.value = f"{group_name}"
        group_cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        group_cell.style = 'Headline 1'

        current_row += 2

        # Headers
        for col_num, header in enumerate(results_headers, start=1):
            header_cell = results_worksheet.cell(
                row=current_row, column=col_num)
            header_cell.value = header
            header_cell.style = 'Headline 3'
        current_row += 1

        # Data
        for result in group_results:
            rank_cell = results_worksheet.cell(
                row=current_row, column=1, value=result.rank)
            athlete_cell = results_worksheet.cell(
                row=current_row, column=2, value=f"{result.athlete.first_name} {result.athlete.last_name}"
            )
            heat_time_cell = results_worksheet.cell(row=current_row, column=3,
                                                    value=format_time(result.heat_time))

            if result.rank == 1:
                rank_cell.fill = gold_fill
                athlete_cell.fill = gold_fill
                heat_time_cell.fill = gold_fill
            elif result.rank == 2:
                rank_cell.fill = silver_fill
                athlete_cell.fill = silver_fill
                heat_time_cell.fill = silver_fill
            elif result.rank == 3:
                rank_cell.fill = bronze_fill
                athlete_cell.fill = bronze_fill
                heat_time_cell.fill = bronze_fill

            current_row += 1

        # Adjust column widths
        for col_num in range(1, len(results_headers) + 1):
            col_letter = get_column_letter(col_num)
            results_worksheet.column_dimensions[col_letter].width = 20

    # Save workbook to an in-memory file
    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)
    return file_buffer


@extend_schema(tags=['Download'], request=GroupListSerializer,)
class DownloadEventResults(APIView):

    @extend_schema(

        summary='Send Binary response with the contents of a xlsx file with the event results',
        responses={
            200: OpenApiResponse(description='Binary Response for xlsx file created'),
            404: OpenApiResponse(description='Resource not found')
        }
    )
    def post(self, request, event_id):

        # Get event instance
        event_instance = MeetEvent.objects.get(id=event_id)

        swim_meet_instance = event_instance.swim_meet

        swim_meet_details = {
            'name': swim_meet_instance.name,
            'date': swim_meet_instance.date.strftime('%m/%d/%Y'),
            'location': swim_meet_instance.site.name
        }

        general_results = self.get_queryset(event_instance)

        serializer = GroupsListSerializer(data=request.data)
        if serializer.is_valid():
            group_ids = serializer.validated_data['group_ids']
            if not group_ids:
                group_ids = []
            print(group_ids)
            results_by_group = {"General Results": general_results}
            for group_id in group_ids:
                group_results = self.get_queryset(event_instance, group_id)
                if group_results:
                    group_instance = Group.objects.get(id=group_id)
                    group_name = group_instance.name
                    results_by_group[group_name] = group_results
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Generate Excel file
        file = create_excel_file_for_event_results(swim_meet_details,
                                                   event_instance.name, results_by_group)
        file_name = f"{event_instance.name}_Results.xlsx"

        # Create HTTP response
        response = HttpResponse(
            file,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        return response

    def get_queryset(self, event_instance, group_id=None):
        queryset = Heat.objects.filter(event=event_instance)
        if group_id is not None:
            queryset = filter_by_group(group_id, queryset, False)

        results = list(queryset.order_by(
            'heat_time', 'athlete__last_name', 'athlete__first_name'))
        if results:
            current_rank = 1
            previous_time = -1
            for i, result in enumerate(results):
                if result.heat_time in [timedelta(days=300), timedelta(days=400), None]:
                    result.rank = ""
                elif result.heat_time == previous_time:
                    result.rank = current_rank
                else:
                    current_rank = i + 1
                    result.rank = current_rank
                previous_time = result.heat_time
        return results
