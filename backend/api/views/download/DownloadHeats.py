from io import BytesIO
from api.models import Heat, MeetEvent, SwimMeet
from django.http import HttpResponse
from drf_spectacular.utils import OpenApiResponse, extend_schema
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from rest_framework.views import APIView
from datetime import timedelta
from api.utils.HelperFunctions import format_time


def create_excel_file_for_heats_details(swim_meet_details, event_data):

    workbook = Workbook()
    # ========================= By Heats Worksheet =========================
    heats_worksheet = workbook.active
    heats_worksheet.title = "By Heats"
    heats_headers = ["Lane", "Athlete", "Seed Time", "Heat Time"]
    current_row = 1

    heats_worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(heats_headers))
    swim_meet_cell = heats_worksheet.cell(row=1, column=1)
    swim_meet_cell.value = f"{swim_meet_details['name']}\n{swim_meet_details['date']}\n{swim_meet_details['location']}"
    swim_meet_cell.style = 'Title'
    swim_meet_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    current_row += 3
    for event in event_data:
        event_cell = heats_worksheet.cell(row=current_row, column=1)
        event_cell.value = event["event_name"]
        heats_worksheet.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=len(heats_headers))
        event_cell.style = 'Headline 1'
        current_row += 2

        if event["heats"]:
            for heat in event["heats"]:
                heat_name = heat["heat_name"]
                lanes = heat["lanes"]
                heat_cell = heats_worksheet.cell(row=current_row, column=1)
                heat_cell.value = heat_name
                heats_worksheet.merge_cells(
                    start_row=current_row, start_column=1, end_row=current_row, end_column=len(heats_headers))
                heat_cell.style = 'Headline 2'

                current_row += 1

                # Headers
                for col_num, header in enumerate(heats_headers, start=1):
                    header_cell = heats_worksheet.cell(
                        row=current_row, column=col_num)
                    header_cell.value = header
                    header_cell.style = 'Headline 3'

                current_row += 1

                # Data
                for lane in lanes:
                    heats_worksheet.cell(
                        row=current_row, column=1, value=lane["lane_num"])
                    heats_worksheet.cell(
                        row=current_row, column=2, value=lane["athlete_full_name"])
                    heats_worksheet.cell(
                        row=current_row, column=3, value=lane["seed_time"])
                    heats_worksheet.cell(
                        row=current_row, column=4, value=lane["heat_time"])
                    current_row += 1
                current_row += 2
        else:

            heats_worksheet.cell(row=current_row, column=1,
                                 value="This event has no heats yet.")
            current_row += 2

    for col_num in range(1, len(heats_headers) + 1):
        col_letter = get_column_letter(col_num)
        heats_worksheet.column_dimensions[col_letter].width = 15

    # ========================= By Lanes Worksheet =========================
    lanes_worksheet = workbook.create_sheet(title="By Lanes")
    lanes_headers = ["Heat ", "Heat Time"]
    current_row = 1

    lanes_worksheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=len(heats_headers))
    swim_meet_cell = lanes_worksheet.cell(row=1, column=1)
    swim_meet_cell.value = f"{swim_meet_details['name']}\n{swim_meet_details['date']}\n{swim_meet_details['location']}"
    swim_meet_cell.style = 'Title'
    swim_meet_cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)

    current_row += 3

    for event in event_data:
        event_cell = lanes_worksheet.cell(row=current_row, column=1)
        event_cell.value = event["event_name"]
        lanes_worksheet.merge_cells(
            start_row=current_row, start_column=1, end_row=current_row, end_column=len(lanes_headers))
        event_cell.alignment = Alignment(
            horizontal='center', vertical='center')
        event_cell.style = 'Headline 1'

        current_row += 2

        if event["heats"]:
            for lane in event["lanes"]:
                lane_name = lane["lane_name"]
                heats = lane["heats"]

                lane_cell = lanes_worksheet.cell(row=current_row, column=1)
                lane_cell.value = lane_name
                lanes_worksheet.merge_cells(
                    start_row=current_row, start_column=1, end_row=current_row, end_column=len(lanes_headers))
                current_row += 1
                lane_cell.style = 'Headline 2'

                # Headers
                header_cell = lanes_worksheet.cell(
                    row=current_row, column=2)
                header_cell.value = lanes_headers[1]
                header_cell.style = 'Headline 3'

                # Data
                for heat in heats:
                    header_cell = lanes_worksheet.cell(
                        row=current_row, column=1)
                    header_cell.value = lanes_headers[0] + \
                        str(heat["num_heat"])
                    header_cell.style = 'Headline 3'
                    current_row += 1

                    lanes_worksheet.cell(
                        row=current_row, column=1, value=heat["athlete_full_name"])
                    lanes_worksheet.cell(
                        row=current_row, column=2, value=heat["heat_time"])
                    current_row += 2

                current_row += 1

        else:

            lanes_worksheet.cell(row=current_row, column=1,
                                 value="This event has no heats yet.")
            current_row += 2

    for col_num in range(1, len(lanes_headers) + 1):
        col_letter = get_column_letter(col_num)
        lanes_worksheet.column_dimensions[col_letter].width = 15

    # Save workbook to an in-memory file
    file_buffer = BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)

    return file_buffer


@extend_schema(tags=['Download'])
class DownloadAllHeatsByEvent(APIView):

    @extend_schema(
        summary='Send Binary response with the contents of a xlsx file with the heat details for the given event',
        responses={
            200: OpenApiResponse(description='Binary Response for xlsx file created'),
            404: OpenApiResponse(description='Resource not found')
        }
    )
    def post(self, request, event_id):

        event_data = []

        # Get event instance
        event_instance = MeetEvent.objects.get(id=event_id)
        event_name = event_instance.name

        # Get number of lanes on the event
        swim_meet_instance = event_instance.swim_meet

        swim_meet_details = {
            'name': swim_meet_instance.name,
            'date': swim_meet_instance.date.strftime('%m/%d/%Y'),
            'location': swim_meet_instance.site.name
        }

        num_lanes = swim_meet_instance.site.num_lanes

        max_num_heat = event_instance.total_num_heats

        by_heats_data = []
        for heat_num in range(1, max_num_heat+1):
            lanes_data = []
            for lane_num in range(1, num_lanes + 1):
                lane = Heat.objects.filter(
                    event_id=event_id, num_heat=heat_num, lane_num=lane_num).first()
                seed_time = format_time(lane.seed_time) if lane else ""
                heat_time = format_time(lane.heat_time) if lane else ""
                lanes_data.append(
                    {
                        "lane_num": lane_num,
                        "athlete_full_name": lane.athlete.full_name if lane and lane.athlete else None,
                        "seed_time": seed_time,
                        "heat_time": heat_time,
                    }
                )

            by_heats_data.append({
                "id": heat_num,
                "heat_name": f"Heat {heat_num} of {max_num_heat}",
                "lanes": lanes_data
            })
        by_lanes_data = []

        for lane_num in range(1, num_lanes+1):
            heats_data = []
            for heat_num in range(1, max_num_heat + 1):
                heat = Heat.objects.filter(
                    event_id=event_id, lane_num=lane_num, num_heat=heat_num).first()
                heat_time = format_time(heat.heat_time) if heat else ""
                heats_data.append(
                    {
                        "num_heat": heat_num,
                        "athlete_full_name": heat.athlete.full_name if heat and heat.athlete else None,
                        "heat_time": heat_time,
                    }
                )

            by_lanes_data.append({
                "id": lane_num,
                "lane_name": f"Lane {lane_num} of {num_lanes}",
                "heats": heats_data
            })
        event_data.append(
            {"event_name": event_name, "heats": by_heats_data, "lanes": by_lanes_data})

        file = create_excel_file_for_heats_details(
            swim_meet_details, event_data)
        file_name = swim_meet_instance.name + "-" + event_name + ".xlsx"

        # Create HTTP response
        response = HttpResponse(
            file,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename=' + file_name
        return response


@extend_schema(tags=['Download'])
class DownloadAllHeatsByMeet(APIView):

    @extend_schema(
        summary='Send Binary response with the contents of a xlsx file with the heat details for all the events of the given swim meet',
        responses={
            200: OpenApiResponse(description='Binary Response for xlsx file created'),
            404: OpenApiResponse(description='Resource not found')
        }
    )
    def post(self, request, meet_id):
        event_data = []

        # Get number of lanes on the event
        swim_meet_instance = SwimMeet.objects.get(id=meet_id)
        num_lanes = swim_meet_instance.site.num_lanes

        swim_meet_details = {
            'name': swim_meet_instance.name,
            'date': swim_meet_instance.date.strftime('%m/%d/%Y'),
            'location': swim_meet_instance.site.name
        }

        events_instance = MeetEvent.objects.filter(
            swim_meet=meet_id).order_by('num_event')
        for event in events_instance:
            event_name = event.name

            max_num_heat = event.total_num_heats

            by_heats_data = []
            by_lanes_data = []
            if max_num_heat:
                for heat_num in range(1, max_num_heat + 1):
                    lanes_data = []
                    for lane_num in range(1, num_lanes + 1):
                        lane = Heat.objects.filter(
                            event_id=event.id, num_heat=heat_num, lane_num=lane_num
                        ).first()
                        seed_time = format_time(
                            lane.seed_time) if lane else ""
                        heat_time = format_time(
                            lane.heat_time) if lane else ""
                        lanes_data.append(
                            {
                                "lane_num": lane_num,
                                "athlete_full_name": lane.athlete.full_name if lane and lane.athlete else None,
                                "seed_time": seed_time,
                                "heat_time": heat_time,
                            }
                        )

                    by_heats_data.append(
                        {"id": heat_num, "heat_name": f"Heat {heat_num} of {max_num_heat}",
                            "lanes": lanes_data}
                    )

                for lane_num in range(1, num_lanes + 1):
                    heats_data = []
                    for heat_num in range(1, max_num_heat + 1):
                        heat = Heat.objects.filter(
                            event_id=event.id, lane_num=lane_num, num_heat=heat_num
                        ).first()
                        heat_time = format_time(
                            heat.heat_time) if heat else ""
                        heats_data.append(
                            {
                                "num_heat": heat_num,
                                "athlete_full_name": heat.athlete.full_name if heat and heat.athlete else None,
                                "heat_time": heat_time,
                            }
                        )
                    by_lanes_data.append(
                        {"id": lane_num, "lane_name": f"Lane {lane_num} of {num_lanes}",
                            "heats": heats_data}
                    )

            event_data.append(
                {
                    "event_name": event_name,
                    "heats": by_heats_data if by_heats_data else None,
                    "lanes": by_lanes_data if by_lanes_data else None,
                }
            )

        file = create_excel_file_for_heats_details(
            swim_meet_details, event_data)
        file_name = swim_meet_instance.name + "- All Event Heat Details.xlsx"

        # Create HTTP response
        response = HttpResponse(
            file,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename=' + file_name
        return response
